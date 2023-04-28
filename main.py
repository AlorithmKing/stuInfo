import sqlite3
import sys

import openpyxl
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QVBoxLayout, QPushButton, \
    QWidget, QFileDialog
from PySide6.QtWidgets import QMessageBox

from Modules.window import *
from Chart import ChartWindow
# 读取ini文件
import configparser

config = configparser.ConfigParser()
config.read('./Modules/usr.ini', encoding='utf-8')
username = config.get("usr", "username")
password = config.get("usr", "password")


class MainWindow(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)

        # SET AS GLOBAL WIDGETS
        # ///////////////////////////////////////////////////////////////

        self.secondWindow = None
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.loginflag = False
        global widgets
        widgets = self.ui
        # self.ui.info_table.setHorizontalHeaderLabels(
        #     ["姓名", "学号", "性别", "专业", "C语言成绩", "Python成绩", "英语成绩"])
        # Signals
        widgets.btn_log.clicked.connect(self.login)
        widgets.btn_exit.clicked.connect(self.exit)
        widgets.btn_Login.clicked.connect(self.buttonClick)
        widgets.btn_PersonalInfo.clicked.connect(self.buttonClick)
        widgets.btn_PersonScores.clicked.connect(self.buttonClick)
        widgets.btn_import.clicked.connect(self.import_Infodata)
        widgets.Score_import.clicked.connect(self.import_Scoredata)
        widgets.btn_insert.clicked.connect(self.insert_selected_row)
        widgets.btn_delete.clicked.connect(self.delete_selected_row)
        widgets.btn_search.clicked.connect(self.show_table_by_assign_data)
        widgets.btn_export.clicked.connect(self.export_data_to_Infodatabase)
        widgets.Score_export.clicked.connect(self.export_data_to_Scoredatabase)
        widgets.btn_sort.clicked.connect(self.sort_by_Scores)
        widgets.btn_show.clicked.connect(self.show_chart)

    def login(self):
        usrname = widgets.line_usr.text()
        pwd = widgets.line_pwd.text()
        message = QMessageBox()
        if usrname == username and pwd == password:
            self.loginflag = True
            message.setText("登录成功！")
            message.setWindowTitle("提示")
            message.exec()
        else:

            # 输出提示信息
            message.setText("请输入正确的用户名和密码才能登录！")
            message.setWindowTitle("提示")
            message.exec()

    def exit(self):
        message = QMessageBox()

        if self.loginflag:
            self.loginflag = False
            message.setText("退出成功！")
            message.setWindowTitle("提示")
            message.exec()
        else:
            message.setText("请先登录！")
            message.setWindowTitle("提示")
            message.exec()


# Slots
    def buttonClick(self):
        # GET BUTTON CLICKED
        btn = self.sender()
        btnName = btn.objectName()
        widgets.line_usr.text()

        # SHOW HOME PAGE
        if btnName == "btn_Login":
            widgets.stackedWidget.setCurrentIndex(2)
            print("Login")
            # UIFunctions.resetStyle(self, btnName)
            # btn.setStyleSheet(UIFunctions.selectMenu(btn.styleSheet()))
        if self.loginflag:
            if btnName == "btn_PersonalInfo":
                widgets.stackedWidget.setCurrentIndex(0)
                # table column setting

                print("PersonalInfo")
            if btnName == "btn_PersonScores":
                widgets.stackedWidget.setCurrentIndex(1)
            # UIFunctions.resetStyle(self, btnName)
            # btn.setStyleSheet(UIFunctions.selectMenu(btn.styleSheet()))
            print("PersonScores")
        else:
            message = QMessageBox()
            # 输出提示信息
            message.setText("请先登入！")
            message.setWindowTitle("警告")
            message.exec()


    def import_Infodata(self):
        file_name, _ = QFileDialog.getOpenFileName(None, "选择Excel文件", "", "Excel Files (*.xlsx *.xls)")
        if file_name:
            workbook = openpyxl.load_workbook(file_name)
            worksheet = workbook.active

            row_count = worksheet.max_row
            column_count = worksheet.max_column

            self.ui.info_table.setRowCount(row_count - 1)
            self.ui.info_table.setColumnCount(column_count)

            for i, row in enumerate(worksheet.iter_rows(min_row=2)):
                for j, cell in enumerate(row):
                    self.ui.info_table.setItem(i, j, QTableWidgetItem(str(cell.value)))

            headers = [cell.value for cell in worksheet[1]]
            self.ui.info_table.setHorizontalHeaderLabels(headers)
        # 设置表格自适应
        self.ui.info_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        # self.ui.info_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch) # 设置行高自适应


    def import_Scoredata(self):
        file_name, _ = QFileDialog.getOpenFileName(None, "选择Excel文件", "", "Excel Files (*.xlsx *.xls)")
        if file_name:
            workbook = openpyxl.load_workbook(file_name)
            worksheet = workbook.active

            row_count = worksheet.max_row
            column_count = worksheet.max_column

            self.ui.score_table.setRowCount(row_count - 1)
            self.ui.score_table.setColumnCount(column_count)

            for i, row in enumerate(worksheet.iter_rows(min_row=2)):
                for j, cell in enumerate(row):
                    self.ui.score_table.setItem(i, j, QTableWidgetItem(str(cell.value)))

            headers = [cell.value for cell in worksheet[1]]
            self.ui.score_table.setHorizontalHeaderLabels(headers)
        # 设置表格自适应
        self.ui.score_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        # self.ui.info_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch) # 设置行高自适应


    def insert_selected_row(self):
        selected_row = self.ui.info_table.currentRow() + 1  # 插入到当前行的下一行
        self.ui.info_table.insertRow(selected_row)


    def delete_selected_row(self):
        selected_row = self.ui.info_table.currentRow()
        if selected_row != -1:
            self.ui.info_table.removeRow(selected_row)


    # 按指定信息显示table的行
    def show_table_by_assign_data(self):
        option = self.ui.cbx_search.currentText()
        info = self.ui.line_search.text()

        if option == "姓名":
            self.search_by_name(info)
        elif option == "学号":
            self.search_by_id(info)
        elif option == "专业":
            self.search_by_major(info)
        else:
            self.search_default()


    def search_by_name(self, name):
        row_count = self.ui.info_table.rowCount()
        for i in range(row_count):
            if self.ui.info_table.item(i, 0).text() == name:
                self.ui.info_table.showRow(i)
            else:
                self.ui.info_table.hideRow(i)


    def search_by_id(self, id):
        row_count = self.ui.info_table.rowCount()
        for i in range(row_count):
            if self.ui.info_table.item(i, 1).text() == id:
                self.ui.info_table.showRow(i)
            else:
                self.ui.info_table.hideRow(i)


    def search_by_major(self, major):
        row_count = self.ui.info_table.rowCount()
        for i in range(row_count):
            if self.ui.info_table.item(i, 3).text() == major:
                self.ui.info_table.showRow(i)
            else:
                self.ui.info_table.hideRow(i)


    def search_default(self):
        row_count = self.ui.info_table.rowCount()
        for i in range(row_count):
            self.ui.info_table.showRow(i)


    # 导出到数据库
    def export_data_to_Infodatabase(self):
        # 创建数据库连接
        conn = sqlite3.connect('./database/student_infor.db')
        cursor = conn.cursor()

        # 创建表（如果不存在）
        cursor.execute('''CREATE TABLE IF NOT EXISTS students (
                                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                                    name TEXT,
                                    ID_number INTEGER,
                                    gender TEXT,
                                    major TEXT,
                                    grade TEXT,
                                    Access_year INTEGER,
                                    Address TEXT,
                                    Parents TEXT,
                                    contact INTEGER)''')
        row_count = self.ui.info_table.rowCount()
        # 遍历表格中的数据
        for row in range(row_count):
            name = self.ui.info_table.item(row, 0).text()
            id = int(self.ui.info_table.item(row, 1).text())
            gender = self.ui.info_table.item(row, 2).text()
            major = self.ui.info_table.item(row, 3).text()
            grade = self.ui.info_table.item(row, 4).text()
            Access_year = int(self.ui.info_table.item(row, 5).text())
            Address = self.ui.info_table.item(row, 6).text()
            Parents = self.ui.info_table.item(row, 7).text()
            contact = int(self.ui.info_table.item(row, 8).text())
            # 将数据插入到数据库表中
            cursor.execute(
                "INSERT INTO students (name, ID_number, gender, major,grade,Access_year,Address,Parents,contact) "
                "VALUES (?, ?, ?, ?,?, ?, ?, ?,?)",
                (name, id, gender, major, grade, Access_year, Address, Parents, contact))

        # 提交更改并关闭连接
        conn.commit()
        conn.close()


    def export_data_to_Scoredatabase(self):
        # 创建数据库连接
        conn = sqlite3.connect('./database/student_Score.db')
        cursor = conn.cursor()

        # 创建表（如果不存在）
        cursor.execute('''CREATE TABLE IF NOT EXISTS students (
                                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                                      name TEXT,
                                      ID_number INTEGER,
                                      gender TEXT,
                                      major TEXT,
                                      C_Score INTEGER,
                                      Python_Score INTEGER,
                                      Eng_Score INTEGER)''')
        row_count = self.ui.info_table.rowCount()
        # 遍历表格中的数据
        for row in range(row_count):
            name = self.ui.info_table.item(row, 0).text()
            id = int(self.ui.info_table.item(row, 1).text())
            gender = self.ui.info_table.item(row, 2).text()
            major = self.ui.info_table.item(row, 3).text()
            C_Score = self.ui.info_table.item(row, 4).text()
            Python_Score = int(self.ui.info_table.item(row, 5).text())
            Eng_Score = self.ui.info_table.item(row, 6).text()

            # 将数据插入到数据库表中
            cursor.execute(
                "INSERT INTO students (name, ID_number, gender, major,C_Score,Python_Score,Eng_Score) "
                "VALUES (?, ?, ?, ?,?, ?, ?)",
                (name, id, gender, major, C_Score, Python_Score, Eng_Score))

        # 提交更改并关闭连接
        conn.commit()
        conn.close()


    def sort_by_Scores(self):
        sel = self.ui.cbxscoremode.currentText()
        sortmode = self.ui.cbx_score.currentText()
        if sel == "升序":
            if sortmode == "C语言":
                self.ui.score_table.sortByColumn(4, Qt.SortOrder.AscendingOrder)
            elif sortmode == "Python":
                self.ui.score_table.sortByColumn(5, Qt.SortOrder.AscendingOrder)
            else:
                self.ui.score_table.sortByColumn(6, Qt.SortOrder.AscendingOrder)

        else:
            if sortmode == "C语言":
                self.ui.score_table.sortByColumn(4, Qt.SortOrder.DescendingOrder)
            elif sortmode == "Python":
                self.ui.score_table.sortByColumn(5, Qt.SortOrder.DescendingOrder)
            else:
                self.ui.score_table.sortByColumn(6, Qt.SortOrder.DescendingOrder)


    def show_chart(self):
        self.secondWindow = ChartWindow()
        self.secondWindow.show()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec())
    # print(username)
